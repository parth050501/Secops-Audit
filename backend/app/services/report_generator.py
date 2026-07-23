"""
Audit report generation — PDF (formal package) and Excel (control matrix).
Pure-Python, no external services. Generates downloadable bytes.
"""
import io
from datetime import datetime


# ── PDF GENERATION (reportlab) ──────────────────────────────────────────────────
def generate_pdf_report(report: dict) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            topMargin=20*mm, bottomMargin=20*mm,
                            leftMargin=18*mm, rightMargin=18*mm)
    styles = getSampleStyleSheet()
    elems = []

    DARK = colors.HexColor("#0f172a")
    TEAL = colors.HexColor("#14b8a6")
    GRAY = colors.HexColor("#64748b")
    LIGHT = colors.HexColor("#f1f5f9")

    h1 = ParagraphStyle("h1", parent=styles["Heading1"], textColor=DARK, fontSize=20, spaceAfter=4)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], textColor=DARK, fontSize=13, spaceBefore=14, spaceAfter=6)
    body = ParagraphStyle("body", parent=styles["Normal"], fontSize=9.5, leading=14, textColor=colors.HexColor("#334155"))
    small = ParagraphStyle("small", parent=styles["Normal"], fontSize=8, textColor=GRAY)

    tenant = report.get("tenant", {})
    summary = report.get("summary", {})

    # ── Branded logo header ──
    # A drawn brand mark (teal shield tile + wordmark) — reliable in-PDF, no
    # external image file needed, so it always renders in the download.
    logo_style = ParagraphStyle("logo", parent=styles["Normal"], fontSize=15,
                                textColor=colors.white, leading=18)
    brand_tbl = Table([[
        Paragraph("<b>◆</b>", ParagraphStyle("mark", parent=styles["Normal"],
                  fontSize=18, textColor=colors.white, alignment=1)),
        Paragraph("<b>GRCBridge</b> &nbsp;<font size=9 color='#99f6e4'>Governance &amp; Compliance Platform</font>", logo_style),
    ]], colWidths=[12*mm, 153*mm])
    brand_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), DARK),
        ("BACKGROUND", (0,0), (0,0), TEAL),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING", (0,0), (-1,-1), 8), ("BOTTOMPADDING", (0,0), (-1,-1), 8),
        ("LEFTPADDING", (0,0), (0,0), 6), ("LEFTPADDING", (1,0), (1,0), 10),
    ]))
    elems.append(brand_tbl)
    elems.append(Spacer(1, 12))

    # ── Cover / header ──
    elems.append(Paragraph("Compliance Audit Report", h1))
    elems.append(Paragraph(f"{tenant.get('name','Organization')} — {tenant.get('framework_name','Framework')}", body))
    elems.append(Paragraph(f"Industry: {tenant.get('industry','').title()} &nbsp;|&nbsp; Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", small))
    elems.append(Spacer(1, 10))

    # ── Score banner ──
    score = report.get("score", 0)
    score_color = colors.HexColor("#10b981" if score >= 80 else "#f59e0b" if score >= 60 else "#ef4444")
    score_tbl = Table([[
        Paragraph(f"<font size=32 color='{score_color.hexval()}'><b>{score}</b></font><font size=12 color='#64748b'>/100</font>", body),
        Paragraph(f"<b>{summary.get('critical',0)}</b> Critical &nbsp;&nbsp; <b>{summary.get('high',0)}</b> High &nbsp;&nbsp; "
                  f"<b>{summary.get('open_tickets',0)}</b> Open Tickets &nbsp;&nbsp; <b>{summary.get('resolved',0)}</b> Resolved", body),
    ]], colWidths=[45*mm, 120*mm])
    score_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), LIGHT),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING", (0,0), (-1,-1), 12), ("BOTTOMPADDING", (0,0), (-1,-1), 12),
        ("LEFTPADDING", (0,0), (-1,-1), 14),
        ("ROUNDEDCORNERS", [6,6,6,6]),
    ]))
    elems.append(score_tbl)
    elems.append(Spacer(1, 6))

    # ── Executive summary ──
    elems.append(Paragraph("Executive Summary", h2))
    for para in (report.get("ai_summary","") or "").split("\n\n"):
        if para.strip():
            elems.append(Paragraph(para.strip(), body))
            elems.append(Spacer(1, 4))

    # ── Control status ──
    elems.append(Paragraph("Control Assessment", h2))
    controls = report.get("controls", [])
    ctrl_rows = [["Control", "Title", "Status", "Findings"]]
    for c in controls:
        ctrl_rows.append([
            c.get("id",""), _truncate(c.get("title",""), 55),
            c.get("status","").upper(), str(c.get("open_findings",0)),
        ])
    ctrl_tbl = Table(ctrl_rows, colWidths=[24*mm, 95*mm, 24*mm, 20*mm], repeatRows=1)
    ctrl_style = [
        ("BACKGROUND", (0,0), (-1,0), DARK),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("FONTSIZE", (0,0), (-1,-1), 8),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("GRID", (0,0), (-1,-1), 0.4, colors.HexColor("#e2e8f0")),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#f8fafc")]),
        ("TOPPADDING", (0,0), (-1,-1), 4), ("BOTTOMPADDING", (0,0), (-1,-1), 4),
        ("LEFTPADDING", (0,0), (-1,-1), 6),
    ]
    for i, c in enumerate(controls, start=1):
        if c.get("status") == "failing":
            ctrl_style.append(("TEXTCOLOR", (2,i), (2,i), colors.HexColor("#ef4444")))
        else:
            ctrl_style.append(("TEXTCOLOR", (2,i), (2,i), colors.HexColor("#10b981")))
    ctrl_tbl.setStyle(TableStyle(ctrl_style))
    elems.append(ctrl_tbl)

    # ── Custom policies (if any) ──
    policies = report.get("custom_policies", [])
    if policies:
        elems.append(Paragraph("Custom Company Policies", h2))
        pol_rows = [["Policy", "Title", "Mode", "Status"]]
        for p in policies:
            pol_rows.append([p.get("policy_id") or f"#{p.get('id')}", _truncate(p.get("title",""), 50),
                             p.get("eval_mode",""), p.get("status","").upper()])
        pol_tbl = Table(pol_rows, colWidths=[26*mm, 90*mm, 24*mm, 23*mm], repeatRows=1)
        pol_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), DARK), ("TEXTCOLOR", (0,0), (-1,0), colors.white),
            ("FONTSIZE", (0,0), (-1,-1), 8), ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("GRID", (0,0), (-1,-1), 0.4, colors.HexColor("#e2e8f0")),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#f8fafc")]),
            ("TOPPADDING", (0,0), (-1,-1), 4), ("BOTTOMPADDING", (0,0), (-1,-1), 4),
            ("LEFTPADDING", (0,0), (-1,-1), 6),
        ]))
        elems.append(pol_tbl)

    elems.append(PageBreak())

    # ── Tickets ──
    elems.append(Paragraph("Remediation Tickets", h2))
    tickets = report.get("tickets", [])
    tk_rows = [["Ref", "Title", "Severity", "Status"]]
    for t in tickets:
        tk_rows.append([t.get("ref",""), _truncate(t.get("title",""), 60),
                        t.get("severity","").upper(), t.get("status","").replace("_"," ").title()])
    tk_tbl = Table(tk_rows, colWidths=[24*mm, 95*mm, 22*mm, 22*mm], repeatRows=1)
    tk_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), DARK), ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("FONTSIZE", (0,0), (-1,-1), 8), ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("GRID", (0,0), (-1,-1), 0.4, colors.HexColor("#e2e8f0")),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#f8fafc")]),
        ("TOPPADDING", (0,0), (-1,-1), 4), ("BOTTOMPADDING", (0,0), (-1,-1), 4),
        ("LEFTPADDING", (0,0), (-1,-1), 6),
    ]))
    elems.append(tk_tbl)

    # ── Audit trail ──
    elems.append(Paragraph("Audit Trail (Chain of Custody)", h2))
    trail = report.get("audit_trail", [])[:30]
    tr_rows = [["Timestamp", "User", "Action"]]
    for e in trail:
        ts = e.get("timestamp","")[:19].replace("T"," ")
        tr_rows.append([ts, e.get("user","System"), e.get("action","").replace("_"," ")])
    tr_tbl = Table(tr_rows, colWidths=[42*mm, 45*mm, 76*mm], repeatRows=1)
    tr_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), DARK), ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("FONTSIZE", (0,0), (-1,-1), 7.5), ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("GRID", (0,0), (-1,-1), 0.4, colors.HexColor("#e2e8f0")),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#f8fafc")]),
        ("TOPPADDING", (0,0), (-1,-1), 3), ("BOTTOMPADDING", (0,0), (-1,-1), 3),
        ("LEFTPADDING", (0,0), (-1,-1), 6),
    ]))
    elems.append(tr_tbl)

    elems.append(Spacer(1, 14))
    elems.append(Paragraph("Generated by GRCBridge — Human-in-the-Loop Governance Platform. "
                           "This report reflects continuous monitoring data and is intended for audit purposes.", small))

    doc.build(elems)
    buf.seek(0)
    return buf.read()


# ── EXCEL GENERATION (openpyxl) ─────────────────────────────────────────────────
def generate_excel_report(report: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    tenant = report.get("tenant", {})

    DARK = "0F172A"; TEAL = "14B8A6"; RED = "EF4444"; GREEN = "10B981"
    HEAD = PatternFill("solid", fgColor=DARK)
    head_font = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(ws, row, ncols):
        for c in range(1, ncols+1):
            cell = ws.cell(row=row, column=c)
            cell.fill = HEAD; cell.font = head_font
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # ── Sheet 1: Summary ──
    ws = wb.active; ws.title = "Summary"
    ws.column_dimensions["A"].width = 28; ws.column_dimensions["B"].width = 50
    ws["A1"] = "Compliance Audit Report"; ws["A1"].font = Font(bold=True, size=16, color=DARK)
    rows = [
        ("Organization", tenant.get("name","")),
        ("Industry", tenant.get("industry","").title()),
        ("Framework", tenant.get("framework_name","")),
        ("Generated", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")),
        ("", ""),
        ("Compliance Score", f"{report.get('score',0)}/100"),
        ("Critical Findings", report.get("summary",{}).get("critical",0)),
        ("High Findings", report.get("summary",{}).get("high",0)),
        ("Open Tickets", report.get("summary",{}).get("open_tickets",0)),
        ("Resolved", report.get("summary",{}).get("resolved",0)),
        ("Connected Systems", report.get("summary",{}).get("connectors",0)),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True, color="64748B")
        ws.cell(row=i, column=2, value=v)

    # ── Sheet 2: Control Matrix ──
    ws2 = wb.create_sheet("Control Matrix")
    headers = ["Control ID", "Title", "Category", "Weight", "Status", "Open Findings"]
    widths = [14, 55, 20, 12, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws2.column_dimensions[chr(64+i)].width = w
    for i, h in enumerate(headers, start=1):
        ws2.cell(row=1, column=i, value=h)
    style_header(ws2, 1, len(headers))
    for r, c in enumerate(report.get("controls", []), start=2):
        ws2.cell(row=r, column=1, value=c.get("id"))
        ws2.cell(row=r, column=2, value=c.get("title"))
        ws2.cell(row=r, column=3, value=c.get("category"))
        ws2.cell(row=r, column=4, value=c.get("weight"))
        sc = ws2.cell(row=r, column=5, value=c.get("status","").upper())
        sc.font = Font(bold=True, color=(RED if c.get("status")=="failing" else GREEN))
        ws2.cell(row=r, column=6, value=c.get("open_findings",0))

    # ── Sheet 3: Custom Policies ──
    policies = report.get("custom_policies", [])
    if policies:
        ws3 = wb.create_sheet("Custom Policies")
        ph = ["Policy ID", "Title", "Category", "Mode", "Status", "Last Result"]
        pw = [16, 45, 18, 12, 14, 50]
        for i, w in enumerate(pw, start=1):
            ws3.column_dimensions[chr(64+i)].width = w
        for i, h in enumerate(ph, start=1):
            ws3.cell(row=1, column=i, value=h)
        style_header(ws3, 1, len(ph))
        for r, p in enumerate(policies, start=2):
            ws3.cell(row=r, column=1, value=p.get("policy_id") or f"#{p.get('id')}")
            ws3.cell(row=r, column=2, value=p.get("title"))
            ws3.cell(row=r, column=3, value=p.get("category"))
            ws3.cell(row=r, column=4, value=p.get("eval_mode"))
            sc = ws3.cell(row=r, column=5, value=p.get("status","").upper())
            sc.font = Font(bold=True, color=(RED if p.get("status")=="failing" else GREEN))
            ws3.cell(row=r, column=6, value=p.get("last_result",""))

    # ── Sheet 4: Tickets ──
    ws4 = wb.create_sheet("Tickets")
    th = ["Ref", "Title", "Severity", "Status", "Framework", "Controls"]
    tw = [14, 55, 12, 14, 14, 20]
    for i, w in enumerate(tw, start=1):
        ws4.column_dimensions[chr(64+i)].width = w
    for i, h in enumerate(th, start=1):
        ws4.cell(row=1, column=i, value=h)
    style_header(ws4, 1, len(th))
    for r, t in enumerate(report.get("tickets", []), start=2):
        ws4.cell(row=r, column=1, value=t.get("ref"))
        ws4.cell(row=r, column=2, value=t.get("title"))
        ws4.cell(row=r, column=3, value=t.get("severity","").upper())
        ws4.cell(row=r, column=4, value=t.get("status","").replace("_"," ").title())
        ws4.cell(row=r, column=5, value=(t.get("framework") or "").upper())
        ws4.cell(row=r, column=6, value=", ".join(t.get("control_ids") or []))

    # ── Sheet 5: Audit Trail ──
    ws5 = wb.create_sheet("Audit Trail")
    ah = ["Timestamp", "User", "Action", "Entity"]
    aw = [24, 24, 28, 16]
    for i, w in enumerate(aw, start=1):
        ws5.column_dimensions[chr(64+i)].width = w
    for i, h in enumerate(ah, start=1):
        ws5.cell(row=1, column=i, value=h)
    style_header(ws5, 1, len(ah))
    for r, e in enumerate(report.get("audit_trail", []), start=2):
        ws5.cell(row=r, column=1, value=e.get("timestamp","")[:19].replace("T"," "))
        ws5.cell(row=r, column=2, value=e.get("user","System"))
        ws5.cell(row=r, column=3, value=e.get("action","").replace("_"," "))
        ws5.cell(row=r, column=4, value=e.get("entity_type",""))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _truncate(s: str, n: int) -> str:
    s = s or ""
    return s if len(s) <= n else s[:n-1] + "…"
<<<<<<< HEAD


# ── LEVELED REPORTS (CISO / engineer / auditor) ──────────────────────────────
def generate_leveled_pdf(report: dict) -> bytes:
    """Render an audience-level report (ciso|engineer|auditor) to PDF."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle)
    import io

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=18*mm, bottomMargin=18*mm,
                            leftMargin=16*mm, rightMargin=16*mm)
    styles = getSampleStyleSheet()
    DARK = colors.HexColor("#1F3A5F"); TEAL = colors.HexColor("#0F8B8D")
    GRAY = colors.HexColor("#64748B")
    h1 = ParagraphStyle("h1", parent=styles["Heading1"], textColor=DARK, fontSize=20, spaceAfter=4)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], textColor=DARK, fontSize=13, spaceBefore=14, spaceAfter=6)
    body = ParagraphStyle("body", parent=styles["Normal"], fontSize=9.5, leading=14, textColor=colors.HexColor("#334155"))
    small = ParagraphStyle("small", parent=styles["Normal"], fontSize=8, textColor=GRAY)

    def rc(pct):
        if pct >= 90: return colors.HexColor("#16a34a")
        if pct >= 75: return colors.HexColor("#65a30d")
        if pct >= 50: return colors.HexColor("#d97706")
        if pct >= 30: return colors.HexColor("#ea580c")
        return colors.HexColor("#dc2626")

    e = []
    # header
    e.append(Paragraph("<b>◆ GRCBridge</b> &nbsp;<font size=9 color='#0F8B8D'>Governance &amp; Compliance Platform</font>",
                       ParagraphStyle("logo", parent=styles["Normal"], fontSize=15, textColor=DARK)))
    e.append(Spacer(1, 8))
    e.append(Paragraph(report.get("level_label", "Compliance Report"), h1))
    t = report.get("tenant", {})
    e.append(Paragraph(f"{t.get('name','Organization')} &nbsp;|&nbsp; {str(t.get('industry','')).title()}", body))
    e.append(Paragraph(f"Generated: {report.get('generated_at','')}", small))
    e.append(Spacer(1, 12))

    level = report.get("level")

    if level == "ciso":
        r = report.get("risk", {})
        e.append(Paragraph(f"<b>Overall readiness: <font color='{rc(report.get('overall_readiness',0)).hexval()}'>{report.get('overall_readiness',0)}%</font></b>", h2))
        e.append(Paragraph(r.get("headline",""), body))
        e.append(Paragraph(f"Open risk: <b>{r.get('critical',0)}</b> critical, <b>{r.get('high',0)}</b> high", body))
        e.append(Spacer(1, 8))
        e.append(Paragraph("Framework posture", h2))
        data = [["Framework", "Readiness", "Passing", "Trend"]]
        for f in report.get("frameworks", []):
            delta = f.get("delta")
            trend = "—" if delta is None else (f"+{delta}%" if delta >= 0 else f"{delta}%")
            data.append([f["name"], f"{f['readiness_pct']}%", f"{f['passing']}/{f['total']}", trend])
        tbl = Table(data, colWidths=[70*mm, 30*mm, 30*mm, 30*mm])
        tbl.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),DARK), ("TEXTCOLOR",(0,0),(-1,0),colors.white),
            ("FONTSIZE",(0,0),(-1,-1),9), ("GRID",(0,0),(-1,-1),0.5,colors.HexColor("#E2E8F0")),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white, colors.HexColor("#F8FAFC")]),
            ("PADDING",(0,0),(-1,-1),6),
        ]))
        e.append(tbl)

    elif level == "engineer":
        for f in report.get("frameworks", []):
            e.append(Paragraph(f"{f['name']} — {f['readiness_pct']}% ready, {f['failing_count']} failing", h2))
            if not f["failing_controls"]:
                e.append(Paragraph("No failing controls. ✓", body)); continue
            for c in f["failing_controls"]:
                e.append(Paragraph(f"<b>{c['id']}</b> — {c['title']} <font color='#dc2626'>[{c.get('weight','')}]</font>", body))
                for ft in c.get("findings", [])[:5]:
                    e.append(Paragraph(f"&nbsp;&nbsp;• {ft}", small))
            e.append(Spacer(1, 6))

    elif level == "auditor":
        for f in report.get("frameworks", []):
            ver = f" ({f['version']})" if f.get("version") else ""
            e.append(Paragraph(f"{f['name']}{ver} — {f['readiness_pct']}% ({f['passing']}/{f['total']})", h2))
            data = [["Control", "Title", "Status"]]
            for c in f.get("controls", [])[:60]:
                data.append([c["id"], c["title"][:60], c["status"].upper()])
            tbl = Table(data, colWidths=[28*mm, 110*mm, 25*mm])
            tbl.setStyle(TableStyle([
                ("BACKGROUND",(0,0),(-1,0),DARK), ("TEXTCOLOR",(0,0),(-1,0),colors.white),
                ("FONTSIZE",(0,0),(-1,-1),7.5), ("GRID",(0,0),(-1,-1),0.4,colors.HexColor("#E2E8F0")),
                ("PADDING",(0,0),(-1,-1),3),
            ]))
            e.append(tbl); e.append(Spacer(1, 8))
        ev = report.get("evidence_sources", [])
        if ev:
            e.append(Paragraph("Evidence sources", h2))
            for s in ev:
                e.append(Paragraph(f"• {s['source']} ({s['type']}) — {s['status']}", small))

    e.append(Spacer(1, 16))
    e.append(Paragraph("Generated by GRCBridge — Human-in-the-Loop Governance Platform. "
                       "Scanned checks cover machine-verifiable controls; other controls require human-attested evidence.", small))
    doc.build(e)
    return buf.getvalue()
=======
>>>>>>> e92200eb629e646833d5267d624ea079b89643af
